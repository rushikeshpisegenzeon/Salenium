import openpyxl


#for same data
# file="C:\\Users\\RPise\\Documents\\Sample_Data.xlsx"
#
# workbook=openpyxl.load_workbook(file)
# sheet = workbook["Sheet2"] #alternate for this (workbook.active)  ---get the active sheet from excel
#
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="Rushikesh"
#
# workbook.save(file)

#for multiple data
file="C:\\Users\\RPise\\Documents\\Sample_Data.xlsx"

workbook=openpyxl.load_workbook(file)
sheet = workbook["Sheet3"]

sheet.cell(1, 1).value = 123
sheet.cell(1, 2).value = "Rushikesh"


sheet.cell(2, 1).value = 124
sheet.cell(2, 2).value = "Vivek"

sheet.cell(3, 1).value = 125
sheet.cell(3, 2).value = "Rahul"

workbook.save(file)




